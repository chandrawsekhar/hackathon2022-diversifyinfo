from cgi import print_exception
import http.client
import json
import openpyxl
import datetime

InputExcelFilename = input("Please enter input excel filename: ")

#datatime for output filename
now = datetime.datetime.now()
date = '{}-{}-{}-{}-{}-{}'.format(now.year, now.month, now.day, now.hour, now.minute, now.second)
outputfilename = 'DiversificationInfo' + '_' + date + '_.xlsx'

#https connection to DNB
conn = http.client.HTTPSConnection("plus.dnb.com")
headers = {
'accept': "application/json;charset=utf-8",
'authorization': "Bearer alphanumerictoken",
}

#Input File
wb_obj = openpyxl.load_workbook(InputExcelFilename)
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row

#Output FIle
wb = openpyxl.Workbook()
sheet = wb.active

#excel headers
iteration = 1
c = sheet.cell(row = iteration, column = 1)
c.value = "DUNS"
c = sheet.cell(row = iteration, column = 2)
c.value = "Name"
c = sheet.cell(row = iteration, column = 3)
c.value = "is8ACertifiedBusiness"
c = sheet.cell(row = iteration, column = 4)
c.value = "isWomanOwned"
c = sheet.cell(row = iteration, column = 5)
c.value = "isMinorityOwned"
c = sheet.cell(row = iteration, column = 6)
c.value = "isVeteranOwned"
c = sheet.cell(row = iteration, column = 7)
c.value = "isVietnamVeteranOwned"
c = sheet.cell(row = iteration, column = 8)
c.value = "ownershipPrimaryEthnicityType"
c = sheet.cell(row = iteration, column = 9)
c.value = "isDisadvantagedBusinessEnterprise"
c = sheet.cell(row = iteration, column = 10)
c.value = "isAirportConcessionDisadvantagedBusinessEnterprise"
c = sheet.cell(row = iteration, column = 11)
c.value = "isAlaskanNativeCorporation"
c = sheet.cell(row = iteration, column = 12)
c.value = "isCertifiedSmallBusiness"
c = sheet.cell(row = iteration, column = 13)
c.value = "isDisabledVeteranBusinessEnterprise"
c = sheet.cell(row = iteration, column = 14)
c.value = "isDisadvantagedVeteranBusinessEnterprise"
c = sheet.cell(row = iteration, column = 15)
c.value = "isFederalHUBCertified"
c = sheet.cell(row = iteration, column = 16)
c.value = "isStateHUBCertified"
c = sheet.cell(row = iteration, column = 17)
c.value = "isMinorityBusinessEnterprise"
c = sheet.cell(row = iteration, column = 18)
c.value = "isServiceDisabledVeteranOwned"
c = sheet.cell(row = iteration, column = 19)
c.value = "isVeteranBusinessEnterprise"
c = sheet.cell(row = iteration, column = 20)
c.value = "isWomanOwnedBusinessEnterprise"
c = sheet.cell(row = iteration, column = 21)
c.value = "isWomanOwnedSmallBusiness"
c = sheet.cell(row = iteration, column = 22)
c.value = "isMinorityServingInstitution"
c = sheet.cell(row = iteration, column = 23)
c.value = "isLGBTQOwned"
c = sheet.cell(row = iteration, column = 24)
c.value = "isSmallDisadvantagedBusiness"
c = sheet.cell(row = iteration, column = 25)
c.value = "isLocalDisadvantagedBusinessEnterprise"
c = sheet.cell(row = iteration, column = 26)
c.value = "isDisabledOwned"
c = sheet.cell(row = iteration, column = 27)
c.value = "classificationDetails"
#end of excel headers



for rownum in range(2, max_row + 1 ):
    iteration += 1
    cell_obj = sheet_obj.cell(row = rownum, column = 1)
    dunsNumber = cell_obj.value
    dunsNumber = str(dunsNumber)
    dunsNumber = dunsNumber.strip()
    #conn.request("GET", "https://plus.dnb.com/v1/data/duns/{dunsNumber}?blockIDs=diversityinsight_L3_v1&tradeUp=hq&customerReference=customer%20reference%20text", headers=headers)
    #res = conn.getresponse()
    #resdata = res.read()
    text_file = open("dataBlocks-sampleL3.json", "r")
    try:
        data = json.load(text_file)
        c = sheet.cell(row = iteration, column = 1)
        c.value = dunsNumber
        c = sheet.cell(row = iteration, column = 2)
        c.value = data["organization"]["primaryName"] 
        c = sheet.cell(row = iteration, column = 3)
        c.value = str(data["organization"]["socioEconomicInformation"]["is8ACertifiedBusiness"])
        c = sheet.cell(row = iteration, column = 4)
        c.value = str(data["organization"]["socioEconomicInformation"]["isWomanOwned"])
        c = sheet.cell(row = iteration, column = 5)
        c.value = str(data["organization"]["socioEconomicInformation"]["isMinorityOwned"])
        c = sheet.cell(row = iteration, column = 6)
        c.value = str(data["organization"]["socioEconomicInformation"]["isVeteranOwned"])
        c = sheet.cell(row = iteration, column = 7)
        c.value = str(data["organization"]["socioEconomicInformation"]["isVietnamVeteranOwned"])
        c = sheet.cell(row = iteration, column = 8)
        c.value = data["organization"]["socioEconomicInformation"]["ownershipPrimaryEthnicityType"]["description"]
        c = sheet.cell(row = iteration, column = 9)
        c.value = str(data["organization"]["socioEconomicInformation"]["isDisadvantagedBusinessEnterprise"])
        c = sheet.cell(row = iteration, column = 10)
        c.value = str(data["organization"]["socioEconomicInformation"]["isAirportConcessionDisadvantagedBusinessEnterprise"])
        c = sheet.cell(row = iteration, column = 11)
        c.value = str(data["organization"]["socioEconomicInformation"]["isAlaskanNativeCorporation"])
        c = sheet.cell(row = iteration, column = 12)
        c.value = str(data["organization"]["socioEconomicInformation"]["isCertifiedSmallBusiness"])
        c = sheet.cell(row = iteration, column = 13)
        c.value = str(data["organization"]["socioEconomicInformation"]["isDisabledVeteranBusinessEnterprise"])
        c = sheet.cell(row = iteration, column = 14)
        c.value = str(data["organization"]["socioEconomicInformation"]["isDisadvantagedVeteranBusinessEnterprise"])
        c = sheet.cell(row = iteration, column = 15)
        c.value = str(data["organization"]["socioEconomicInformation"]["isFederalHUBCertified"])
        c = sheet.cell(row = iteration, column = 16)
        c.value = str(data["organization"]["socioEconomicInformation"]["isStateHUBCertified"])
        c = sheet.cell(row = iteration, column = 17)
        c.value = str(data["organization"]["socioEconomicInformation"]["isMinorityBusinessEnterprise"])
        c = sheet.cell(row = iteration, column = 18)
        c.value = str(data["organization"]["socioEconomicInformation"]["isServiceDisabledVeteranOwned"])
        c = sheet.cell(row = iteration, column = 19)
        c.value = str(data["organization"]["socioEconomicInformation"]["isVeteranBusinessEnterprise"])
        c = sheet.cell(row = iteration, column = 20)
        c.value = str(data["organization"]["socioEconomicInformation"]["isWomanOwnedBusinessEnterprise"])
        c = sheet.cell(row = iteration, column = 21)
        c.value = str(data["organization"]["socioEconomicInformation"]["isWomanOwnedSmallBusiness"])
        c = sheet.cell(row = iteration, column = 22)
        c.value = str(data["organization"]["socioEconomicInformation"]["isMinorityServingInstitution"])
        c = sheet.cell(row = iteration, column = 23)
        c.value = str(data["organization"]["socioEconomicInformation"]["isLGBTQOwned"])
        c = sheet.cell(row = iteration, column = 24)
        c.value = str(data["organization"]["socioEconomicInformation"]["isSmallDisadvantagedBusiness"])
        c = sheet.cell(row = iteration, column = 25)
        c.value = str(data["organization"]["socioEconomicInformation"]["isLocalDisadvantagedBusinessEnterprise"])
        c = sheet.cell(row = iteration, column = 26)
        c.value = str(data["organization"]["socioEconomicInformation"]["isDisabledOwned"])
        c = sheet.cell(row = iteration, column = 27)
        c.value = data["organization"]["socioEconomicInformation"]["classificationDetails"][-1]["classification"]["description"]
        text_file.close()
    except:
        print_exception ()
        #print text_file.readlines()
  

wb.save(outputfilename)
wb.close()
wb_obj.close()








